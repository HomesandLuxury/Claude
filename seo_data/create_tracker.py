"""
create_tracker.py — Creates backlinks_tracker.xlsx with full formatting
Run: python create_tracker.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

SHEET_PATH = os.path.join(os.path.dirname(__file__), "backlinks_tracker.xlsx")

TIERS = {
    "Tier 1 - Directory":    ["Google Business Profile","Bing Places","Yelp Business","Foursquare","Hotfrog","Manta","Yellow Pages","Cylex","n49","Tupalo","EZLocal","MapQuest","Brownbook","2FindLocal","Spoke"],
    "Tier 2 - Profile":      ["Crunchbase","LinkedIn Company","AngelList/Wellfound","Product Hunt","About.me","Gravatar","Behance","Medium","Tumblr","WordPress.com","Blogger","Weebly","Wix"],
    "Tier 3 - Q&A":          ["Quora","Reddit r/realestate","Reddit r/luxuryhomes","Reddit r/homebuying","Stack Exchange DIY"],
    "Tier 4 - Article":      ["LinkedIn Articles","Medium Articles","Vocal.media","HubPages","EzineArticles","ArticleBiz","GoArticles","SelfGrowth","Scoop.it","Flipboard"],
    "Tier 5 - Social":       ["Facebook Business","Instagram Business","Twitter/X","Pinterest Business","YouTube Channel","TikTok Business","Snapchat","Telegram Channel"],
    "Tier 6 - Image":        ["Pinterest Pins","Flickr","500px","Imgur","DeviantArt","Pixabay Profile"],
    "Tier 7 - Niche":        ["Houzz","HomeAdvisor","Realty Times","HGTV Community","This Old House","Homes.com","Luxury Portfolio"],
    "Tier 8 - Press":        ["PRLog","OpenPR","PR.com","1888PressRelease","PRBuzz","NewswireToday","i-Newswire","PRFree"],
    "Tier 9 - Document":     ["Slideshare","Scribd","Issuu"],
    "Tier 10 - Blog Comment":["Blog Comment 1","Blog Comment 2","Blog Comment 3","Blog Comment 4","Blog Comment 5"],
}

SUBMISSION_URLS = {
    "Google Business Profile"   : "https://business.google.com",
    "Bing Places"               : "https://www.bingplaces.com",
    "Yelp Business"             : "https://biz.yelp.com",
    "Foursquare"                : "https://business.foursquare.com",
    "Hotfrog"                   : "https://www.hotfrog.com",
    "Manta"                     : "https://www.manta.com",
    "Yellow Pages"              : "https://www.yellowpages.com",
    "Cylex"                     : "https://www.cylex.us.com",
    "n49"                       : "https://www.n49.com",
    "Tupalo"                    : "https://tupalo.com",
    "EZLocal"                   : "https://www.ezlocal.com",
    "MapQuest"                  : "https://www.mapquest.com/add-a-place",
    "Brownbook"                 : "https://www.brownbook.net",
    "2FindLocal"                : "https://www.2findlocal.com",
    "Spoke"                     : "https://www.spoke.com",
    "Crunchbase"                : "https://www.crunchbase.com",
    "LinkedIn Company"          : "https://linkedin.com/company/new",
    "AngelList/Wellfound"       : "https://wellfound.com",
    "Product Hunt"              : "https://www.producthunt.com",
    "About.me"                  : "https://about.me",
    "Gravatar"                  : "https://gravatar.com",
    "Behance"                   : "https://www.behance.net",
    "Medium"                    : "https://medium.com",
    "Tumblr"                    : "https://tumblr.com",
    "WordPress.com"             : "https://wordpress.com",
    "Blogger"                   : "https://blogger.com",
    "Weebly"                    : "https://www.weebly.com",
    "Wix"                       : "https://www.wix.com",
    "Quora"                     : "https://www.quora.com",
    "Facebook Business"         : "https://facebook.com/pages/create",
    "Instagram Business"        : "https://instagram.com",
    "Twitter/X"                 : "https://twitter.com",
    "Pinterest Business"        : "https://pinterest.com/business/create/",
    "YouTube Channel"           : "https://youtube.com",
    "TikTok Business"           : "https://tiktok.com",
    "Flickr"                    : "https://flickr.com",
    "500px"                     : "https://500px.com",
    "Imgur"                     : "https://imgur.com",
    "Houzz"                     : "https://www.houzz.com",
    "HomeAdvisor"               : "https://www.homeadvisor.com",
    "Realty Times"              : "https://realtytimes.com",
    "PRLog"                     : "https://www.prlog.org",
    "OpenPR"                    : "https://www.openpr.com",
    "Slideshare"                : "https://www.slideshare.net",
    "Scribd"                    : "https://www.scribd.com",
    "Issuu"                     : "https://issuu.com",
    "Vocal.media"               : "https://vocal.media",
    "HubPages"                  : "https://hubpages.com",
    "LinkedIn Articles"         : "https://www.linkedin.com",
    "Scoop.it"                  : "https://www.scoop.it",
    "Flipboard"                 : "https://flipboard.com",
}

def hex2rgb(h):
    h = h.lstrip('#')
    return tuple(int(h[i:i+2],16) for i in (0,2,4))

TIER_COLORS = {
    "Tier 1 - Directory"    : "1B4F72",
    "Tier 2 - Profile"      : "154360",
    "Tier 3 - Q&A"          : "1A5276",
    "Tier 4 - Article"      : "0E6655",
    "Tier 5 - Social"       : "117A65",
    "Tier 6 - Image"        : "1E8449",
    "Tier 7 - Niche"        : "7D6608",
    "Tier 8 - Press"        : "784212",
    "Tier 9 - Document"     : "6E2F1A",
    "Tier 10 - Blog Comment": "4A235A",
}

def create_tracker():
    wb = openpyxl.Workbook()

    # ── MAIN TRACKING SHEET ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Backlinks Tracker"

    headers = ["#","Date Added","Platform / Site","Submission URL","Target Page",
               "Anchor Text","Type","DR","Status","Follow","Category","Notes","Verified"]

    hdr_fill = PatternFill("solid", fgColor="1B3A5C")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    col_widths = [5,13,28,45,38,28,16,8,14,12,22,30,13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Pre-populate all 80+ sites
    row = 2
    for tier, sites in TIERS.items():
        tier_color = TIER_COLORS.get(tier, "444444")
        for site in sites:
            fill = PatternFill("solid", fgColor="F0F4F8" if row % 2 == 0 else "FFFFFF")
            sub_url = SUBMISSION_URLS.get(site, "")
            data = [row-1, "", site, sub_url, "https://homesandluxury.com",
                    "HomesAndLuxury", "", "", "To Do", "Dofollow", tier, "", ""]
            for c, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.fill = fill
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            # Color the tier cell
            tier_cell = ws.cell(row=row, column=11)
            tier_cell.font = Font(name="Calibri", size=10, color=tier_color, bold=True)
            row += 1

    # Add data validation for Status column
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list",
                        formula1='"To Do,Submitted,Live,Rejected,Lost"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.sqref = f"I2:I{row}"

    # ── SUMMARY DASHBOARD ────────────────────────────────────────────
    ws2 = wb.create_sheet("Dashboard")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    title_cell = ws2["A1"]
    title_cell.value = "HomesAndLuxury.com — Backlink Dashboard"
    title_cell.font = Font(bold=True, size=16, name="Calibri", color="1B3A5C")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 35

    ws2["A2"].value = f"Last Updated: {datetime.now().strftime('%d %b %Y')}"
    ws2["A2"].font = Font(size=10, color="888888", name="Calibri", italic=True)

    metrics = [
        ("A4","Total Sites in List",      "=COUNTA('Backlinks Tracker'!C2:C500)"),
        ("A5","Submitted or Live",         "=COUNTIF('Backlinks Tracker'!I2:I500,\"Submitted\")+COUNTIF('Backlinks Tracker'!I2:I500,\"Live\")"),
        ("A6","Live Links",                "=COUNTIF('Backlinks Tracker'!I2:I500,\"Live\")"),
        ("A7","To Do",                     "=COUNTIF('Backlinks Tracker'!I2:I500,\"To Do\")"),
        ("A8","Dofollow",                  "=COUNTIF('Backlinks Tracker'!J2:J500,\"Dofollow\")"),
        ("A9","Nofollow",                  "=COUNTIF('Backlinks Tracker'!J2:J500,\"Nofollow\")"),
        ("A10","Unique Domains",           "=SUMPRODUCT(1/COUNTIF('Backlinks Tracker'!C2:C100,'Backlinks Tracker'!C2:C100))"),
    ]

    label_font  = Font(bold=True, name="Calibri", size=11, color="1B3A5C")
    value_font  = Font(name="Calibri", size=13, bold=True, color="2C3E50")
    box_fill    = PatternFill("solid", fgColor="EBF5FB")

    for ref, label, formula in metrics:
        row_n = int(ref[1:])
        ws2[ref].value = label
        ws2[ref].font  = label_font
        ws2[ref].alignment = Alignment(vertical="center")
        ws2.row_dimensions[row_n].height = 22
        val_ref = ref.replace("A","B")
        ws2[val_ref].value = formula
        ws2[val_ref].font  = value_font
        ws2[val_ref].fill  = box_fill
        ws2[val_ref].alignment = Alignment(horizontal="center", vertical="center")

    # Tier breakdown
    ws2["A12"].value = "Breakdown by Tier"
    ws2["A12"].font = Font(bold=True, size=12, name="Calibri", color="1B3A5C")
    r = 13
    for tier in TIERS:
        ws2.cell(row=r, column=1, value=tier).font = Font(name="Calibri", size=10, color=TIER_COLORS.get(tier,"000000"))
        ws2.cell(row=r, column=2, value=f'=COUNTIF(\'Backlinks Tracker\'!K:K,"{tier}")').font = Font(name="Calibri", size=10, bold=True)
        r += 1

    for c in [1,2,3,4,5,6,7]:
        ws2.column_dimensions[get_column_letter(c)].width = 38 if c == 1 else 15

    # ── SUBMISSION KIT SHEET ─────────────────────────────────────────
    ws3 = wb.create_sheet("Submission Kit")
    kit_data = [
        ("Business Name",       "HomesAndLuxury"),
        ("Website URL",         "https://homesandluxury.com"),
        ("Category",            "Home Decor / Interior Design / Luxury Lifestyle"),
        ("Short Description",   "HomesAndLuxury is a premium home decor and luxury lifestyle blog featuring expert interior design ideas, room styling guides, product reviews, and decor inspiration for homeowners who want beautiful, curated living spaces."),
        ("Long Description",    "HomesAndLuxury.com is your go-to destination for premium home decor ideas, luxury interior design inspiration, and expert lifestyle content. We cover everything from farmhouse and rustic decor to modern luxury living, wall art, gardening, kitchen design, and seasonal decorating. Our editorial team of home decor specialists and interior design consultants publishes in-depth guides, product reviews with honest pros and cons, and trend forecasts to help homeowners create spaces they truly love. Whether you are decorating a small apartment or a luxury home, HomesAndLuxury provides the expert guidance, product recommendations, and visual inspiration you need to bring your vision to life. Explore hundreds of guides covering furniture selection, color palettes, lighting design, outdoor spaces, and the best home decor products available on the market today."),
        ("Keywords",            "home decor ideas, luxury home decor, interior design inspiration, wall decor, farmhouse decor, luxury living, room decorating ideas"),
        ("Email",               "[your contact email]"),
        ("Phone",               "[your phone number]"),
        ("Address",             "[your address or 'Online Only']"),
        ("City",                "[your city]"),
        ("Country",             "[your country]"),
        ("Facebook",            "[your Facebook URL]"),
        ("Instagram",           "[your Instagram URL]"),
        ("Pinterest",           "[your Pinterest URL]"),
        ("Twitter/X",           "[your Twitter URL]"),
        ("YouTube",             "[your YouTube URL]"),
    ]
    ws3["A1"].value = "Submission Kit — Copy-Paste for Every Directory"
    ws3["A1"].font  = Font(bold=True, size=14, name="Calibri", color="1B3A5C")
    ws3.merge_cells("A1:B1")
    for i, (k, v) in enumerate(kit_data, 3):
        ws3.cell(row=i, column=1, value=k).font  = Font(bold=True, name="Calibri", size=10)
        ws3.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.cell(row=i, column=2).font = Font(name="Calibri", size=10)
        ws3.row_dimensions[i].height = 40 if len(v) > 100 else 18
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 90

    wb.save(SHEET_PATH)
    print(f"Backlinks tracker created: {SHEET_PATH}")
    total = sum(len(v) for v in TIERS.values())
    print(f"Pre-loaded {total} backlink sources across {len(TIERS)} tiers")
    print("Open backlinks_tracker.xlsx -> Submission Kit tab to copy-paste business info")

if __name__ == "__main__":
    create_tracker()
